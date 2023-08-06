import json
from copy import deepcopy

from openpyxl import load_workbook

from excel_to_json.worksheet import WorkSheet


class WorkBook:
    def __init__(self):
        self._parsed_sheets = []
        self._json_data = {}

    @property
    def json_datas(self):
        return self._json_data.items()

    def load_file(self, file):
        wb = load_workbook(file)
        for sheetName in filter(lambda e: not e.startswith('#'), wb.sheetnames):
            parsed_sheet = WorkSheet(sheetName)
            parsed_sheet.parse(wb[sheetName])
            self._parsed_sheets.append(parsed_sheet)

        root_sheets = filter(lambda s: '#' not in s.name, self._parsed_sheets)
        for root_sheet in root_sheets:
            json_data = self.combine_parsed_sheet(root_sheet)
            if len(json_data) > 0:
                if root_sheet.is_object:
                    json_data = json_data[0]
                self._json_data[root_sheet.name] = json_data

    def combine_parsed_sheet(self, sheet):
        combined_json_data = deepcopy(sheet.json_data)
        # get directly child sheet
        child_sheets = (s for s in self._parsed_sheets if s.is_child_of(sheet.name))
        # recursive invoke child sheet's combine_parsed_sheet
        for child_sheet in child_sheets:
            child_json_data = self.combine_parsed_sheet(child_sheet)
            # add child's ref to self hook
            for ref_name, ref_rows in child_sheet.refs.items():
                for ref_row_index in ref_rows:
                    if ref_name in sheet.hooks:
                        hook_row_index = sheet.hooks[ref_name]
                        parent_json_data = combined_json_data[hook_row_index]
                        if child_sheet.is_object:
                            parent_json_data[child_sheet.short_name] = child_json_data[ref_row_index]
                        else:
                            if child_sheet.short_name not in parent_json_data:
                                parent_json_data[child_sheet.short_name] = []
                            parent_json_data[child_sheet.short_name].append(child_json_data[ref_row_index])
        return combined_json_data
