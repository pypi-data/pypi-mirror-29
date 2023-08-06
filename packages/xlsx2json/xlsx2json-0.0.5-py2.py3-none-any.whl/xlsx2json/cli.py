import click
import json

from six import string_types
from openpyxl import load_workbook


def parse_hyperlink_formula(value):
    if not isinstance(value, string_types):
        return value, None
    value = value.strip()
    prefix = '=HYPERLINK("'
    suffix = '")'
    if value.startswith(prefix) and value.endswith(suffix):
        value = value[len(prefix):len(value) - len(suffix)]
        href, text = value.split('","', 2)
        return text, href
    else:
        return value, None


@click.command()
@click.argument('input_path')
@click.argument('output_path')
def entry_point(input_path, output_path):
    wb = load_workbook(input_path)
    sheet = wb.worksheets[0]

    rows = []
    headers = None
    for row in sheet.iter_rows():
        if headers is None:
            headers = [('' if cell.value is None else cell.value) for cell in row]
            continue
        values = []
        for cell in row:
            value = {'text': cell.value}
            if cell.hyperlink and cell.hyperlink.target:
                value = {
                    'text': cell.value,
                    'href': cell.hyperlink.target
                }
            else:
                text, href = parse_hyperlink_formula(cell.value)
                value = {'text': text}
                if href:
                    value['href'] = href
            if list(value.keys()) == ['text']:
                value = value['text']
            values.append(value)
        row = dict(zip(headers, values))
        rows.append(row)

    with open(output_path, 'wb') as output_file:
        json_rows = json.dumps(rows, indent=True, ensure_ascii=False).encode()
        output_file.write(json_rows)
