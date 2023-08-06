import click
import json

from openpyxl import load_workbook


@click.command()
@click.argument('input_path')
@click.argument('output_path')
def entry_point(input_path, output_path):
    wb = load_workbook(input_path)
    sheet = wb.worksheets[0]

    rows = []
    headers = None
    for row in sheet.iter_rows():
        if headers is None:
            headers = [cell.value for cell in row]
            continue
        values = []
        for cell in row:
            value = {'text': cell.value}
            if cell.hyperlink and cell.hyperlink.target:
                value['href'] = cell.hyperlink.target
            values.append(value)
        row = dict(zip(headers, values))
        rows.append(row)

    with open(output_path, 'wb') as output_file:
        json_rows = json.dumps(rows, indent=True, ensure_ascii=False).encode()
        output_file.write(json_rows)
